# -*- coding: utf-8 -*-
"""
Actualiza la hoja VARIABLES de Base ONs - Nueva.xlsx.

Entrada:
    DataStorage/variables_mercado.xlsx

Ese archivo tiene que venir generado por:
    GetData/bajar_variables_mercado.py

Actualiza en la hoja VARIABLES:
    A:B -> Fecha / BADLAR
    D:E -> Fecha / TAMAR
    G:H -> Fecha / COM3500

Regla:
    - Para cada bloque, detecta la última fecha cargada.
    - Inserta sólo fechas estrictamente mayores.
    - Copia formato desde la última fila existente de cada bloque.
    - Descarga el .xlsx desde Drive, lo edita con openpyxl y lo vuelve a subir.

Requisitos:
    pip install pandas openpyxl google-api-python-client google-auth
"""

from __future__ import annotations

import io
import sys
import shutil
from copy import copy
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload


# ============================================================
# CONFIG
# ============================================================

DRIVE_FILE_ID = "19Hr6IBST72-D9j2mNTgcBH_-XmI2F9Py"
WORKSHEET_NAME = "VARIABLES"

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name.lower() == "getdata" else Path.cwd()

SERVICE_ACCOUNT_FILE = PROJECT_ROOT / "ServiceAccount" / "service_account.json"

VARIABLES_FILE = PROJECT_ROOT / "DataStorage" / "variables_mercado.xlsx"

LOCAL_DOWNLOAD_FILE = PROJECT_ROOT / "DataStorage" / "Base ONs - Nueva_VARIABLES_DESCARGADA.xlsx"
LOCAL_BACKUP_FILE = PROJECT_ROOT / "DataStorage" / f"backup_Base_ONs_Nueva_VARIABLES_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
LOCAL_UPDATED_FILE = PROJECT_ROOT / "DataStorage" / "Base ONs - Nueva_VARIABLES_ACTUALIZADA.xlsx"

DRY_RUN = False


# ============================================================
# LOG
# ============================================================

def log(msg: str) -> None:
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{ahora} | {msg}")


# ============================================================
# DRIVE API
# ============================================================

def conectar_drive():
    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(f"No encontré credenciales: {SERVICE_ACCOUNT_FILE}")

    scopes = ["https://www.googleapis.com/auth/drive"]

    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=scopes,
    )

    log(f"Service account usada: {creds.service_account_email}")

    service = build("drive", "v3", credentials=creds)
    return service


def descargar_xlsx_drive(service, file_id: str, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)

    meta = service.files().get(
        fileId=file_id,
        fields="id,name,mimeType",
        supportsAllDrives=True,
    ).execute()

    log(f"Archivo Drive encontrado: {meta.get('name')}")
    log(f"MIME type: {meta.get('mimeType')}")

    mime_type = meta.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.spreadsheet":
        raise RuntimeError(
            "El archivo es Google Sheet nativo. Este script está pensado para editar un .xlsx puro."
        )

    request = service.files().get_media(
        fileId=file_id,
        supportsAllDrives=True,
    )

    fh = io.FileIO(destino, "wb")
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            log(f"Descarga: {int(status.progress() * 100)}%")

    fh.close()

    log(f"Archivo descargado en: {destino}")


def subir_xlsx_drive(service, file_id: str, archivo: Path) -> None:
    media = MediaFileUpload(
        str(archivo),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )

    updated = service.files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()

    log(f"Archivo actualizado en Drive OK. ID: {updated.get('id')}")


# ============================================================
# HELPERS
# ============================================================

def parse_fecha(x):
    if x is None:
        return pd.NaT

    if isinstance(x, datetime):
        return pd.Timestamp(x).normalize()

    try:
        if pd.isna(x):
            return pd.NaT
    except Exception:
        pass

    dt = pd.to_datetime(x, errors="coerce", dayfirst=False)

    if pd.isna(dt):
        dt = pd.to_datetime(x, errors="coerce", dayfirst=True)

    if pd.isna(dt):
        return pd.NaT

    return dt.normalize()


def normalizar_numero(x):
    if x is None:
        return pd.NA

    try:
        if pd.isna(x):
            return pd.NA
    except Exception:
        pass

    try:
        return float(x)
    except Exception:
        return pd.NA


def copiar_estilo_celda(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def ultima_fila_real_bloque(ws, fecha_col: int, valor_col: int) -> int:
    """
    Busca última fila real dentro de un bloque de dos columnas.
    Ej:
        A:B, D:E, G:H
    """
    ultima = 1

    max_row = ws.max_row

    for row in range(2, max_row + 1):
        fecha = ws.cell(row=row, column=fecha_col).value
        valor = ws.cell(row=row, column=valor_col).value

        if fecha not in [None, ""] or valor not in [None, ""]:
            ultima = row

    return ultima


def ultima_fecha_bloque(ws, fecha_col: int, valor_col: int) -> pd.Timestamp:
    ultima_fila = ultima_fila_real_bloque(ws, fecha_col, valor_col)

    fechas = []

    for row in range(2, ultima_fila + 1):
        val = ws.cell(row=row, column=fecha_col).value
        dt = parse_fecha(val)

        if not pd.isna(dt):
            fechas.append(dt)

    if not fechas:
        return pd.NaT

    return max(fechas)


def preparar_serie(df: pd.DataFrame, fecha_col: str, valor_col: str) -> pd.DataFrame:
    out = df[[fecha_col, valor_col]].copy()
    out.columns = ["fecha", "valor"]

    out["fecha"] = out["fecha"].apply(parse_fecha)
    out["valor"] = out["valor"].apply(normalizar_numero)

    out = out.dropna(subset=["fecha", "valor"]).copy()
    out = out.sort_values("fecha").drop_duplicates(subset=["fecha"], keep="last").reset_index(drop=True)

    return out


def cargar_variables_mercado() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not VARIABLES_FILE.exists():
        raise FileNotFoundError(f"No encontré el archivo de variables: {VARIABLES_FILE}")

    log(f"Leyendo variables desde: {VARIABLES_FILE}")

    df_badlar_raw = pd.read_excel(VARIABLES_FILE, sheet_name="BADLAR")
    df_tamar_raw = pd.read_excel(VARIABLES_FILE, sheet_name="TAMAR")
    df_com_raw = pd.read_excel(VARIABLES_FILE, sheet_name="COM3500")

    if "fecha" not in df_badlar_raw.columns or "BADLAR_TNA" not in df_badlar_raw.columns:
        raise ValueError(f"La hoja BADLAR no tiene columnas esperadas. Columnas: {list(df_badlar_raw.columns)}")

    if "fecha" not in df_tamar_raw.columns or "TAMAR_TNA" not in df_tamar_raw.columns:
        raise ValueError(f"La hoja TAMAR no tiene columnas esperadas. Columnas: {list(df_tamar_raw.columns)}")

    if "fecha" not in df_com_raw.columns or "COM3500" not in df_com_raw.columns:
        raise ValueError(f"La hoja COM3500 no tiene columnas esperadas. Columnas: {list(df_com_raw.columns)}")

    df_badlar = preparar_serie(df_badlar_raw, "fecha", "BADLAR_TNA")
    df_tamar = preparar_serie(df_tamar_raw, "fecha", "TAMAR_TNA")
    df_com = preparar_serie(df_com_raw, "fecha", "COM3500")

    log(f"BADLAR filas disponibles: {len(df_badlar):,}")
    log(f"TAMAR filas disponibles: {len(df_tamar):,}")
    log(f"COM3500 filas disponibles: {len(df_com):,}")

    return df_badlar, df_tamar, df_com


def aplicar_estilo_bloque(ws, fila_origen: int, fila_destino: int, fecha_col: int, valor_col: int) -> None:
    """
    Copia estilo de la fila anterior dentro del bloque.
    """
    for col in [fecha_col, valor_col]:
        src = ws.cell(row=fila_origen, column=col)
        dst = ws.cell(row=fila_destino, column=col)
        copiar_estilo_celda(src, dst)

    # Aseguramos alineación consistente.
    ws.cell(row=fila_destino, column=fecha_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=fila_destino, column=valor_col).alignment = Alignment(horizontal="center", vertical="center")


def insertar_serie_en_bloque(
    ws,
    nombre: str,
    df_serie: pd.DataFrame,
    fecha_col: int,
    valor_col: int,
) -> int:
    """
    Inserta en un bloque Fecha/Valor sólo las fechas faltantes.
    """
    ultima_fecha = ultima_fecha_bloque(ws, fecha_col, valor_col)
    ultima_fila = ultima_fila_real_bloque(ws, fecha_col, valor_col)

    if pd.isna(ultima_fecha):
        log(f"{nombre}: no encontré fecha previa. Se insertará toda la serie.")
        df_nuevo = df_serie.copy()
    else:
        log(f"{nombre}: última fecha en hoja VARIABLES: {ultima_fecha.strftime('%Y-%m-%d')}")
        df_nuevo = df_serie[df_serie["fecha"] > ultima_fecha].copy()

    df_nuevo = df_nuevo.sort_values("fecha").reset_index(drop=True)

    if df_nuevo.empty:
        log(f"{nombre}: no hay datos nuevos para insertar.")
        return 0

    fila_inicio = ultima_fila + 1

    log(f"{nombre}: filas nuevas a insertar: {len(df_nuevo):,}")
    log(f"{nombre}: rango destino: fila {fila_inicio} en columnas {fecha_col}:{valor_col}")

    for idx, r in df_nuevo.iterrows():
        fila = fila_inicio + idx
        fila_estilo = ultima_fila if idx == 0 else fila - 1

        aplicar_estilo_bloque(
            ws=ws,
            fila_origen=fila_estilo,
            fila_destino=fila,
            fecha_col=fecha_col,
            valor_col=valor_col,
        )

        ws.cell(row=fila, column=fecha_col).value = r["fecha"].to_pydatetime()
        ws.cell(row=fila, column=valor_col).value = float(r["valor"])

        # Preservamos formato del bloque. Si no existe, defaults razonables.
        if not ws.cell(row=fila, column=fecha_col).number_format:
            ws.cell(row=fila, column=fecha_col).number_format = "m/d/yyyy"

        if not ws.cell(row=fila, column=valor_col).number_format:
            ws.cell(row=fila, column=valor_col).number_format = "0.00"

    return len(df_nuevo)


def actualizar_workbook_variables(xlsx_path: Path, df_badlar: pd.DataFrame, df_tamar: pd.DataFrame, df_com: pd.DataFrame) -> int:
    wb = load_workbook(xlsx_path)

    if WORKSHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja {WORKSHEET_NAME} en el archivo.")

    ws = wb[WORKSHEET_NAME]

    # Según estructura vista:
    # A:B -> BADLAR
    # D:E -> TAMAR
    # G:H -> COM3500
    insertados_badlar = insertar_serie_en_bloque(
        ws=ws,
        nombre="BADLAR",
        df_serie=df_badlar,
        fecha_col=1,
        valor_col=2,
    )

    insertados_tamar = insertar_serie_en_bloque(
        ws=ws,
        nombre="TAMAR",
        df_serie=df_tamar,
        fecha_col=4,
        valor_col=5,
    )

    insertados_com = insertar_serie_en_bloque(
        ws=ws,
        nombre="COM3500",
        df_serie=df_com,
        fecha_col=7,
        valor_col=8,
    )

    total = insertados_badlar + insertados_tamar + insertados_com

    if total > 0:
        wb.save(LOCAL_UPDATED_FILE)
        log(f"Workbook actualizado guardado en: {LOCAL_UPDATED_FILE}")
    else:
        log("No hubo datos nuevos en ninguna serie. No se generó archivo actualizado.")

    log(f"Resumen insertados -> BADLAR: {insertados_badlar}, TAMAR: {insertados_tamar}, COM3500: {insertados_com}")

    return total


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    try:
        df_badlar, df_tamar, df_com = cargar_variables_mercado()

        service = conectar_drive()

        descargar_xlsx_drive(service, DRIVE_FILE_ID, LOCAL_DOWNLOAD_FILE)
        shutil.copy2(LOCAL_DOWNLOAD_FILE, LOCAL_BACKUP_FILE)
        log(f"Backup local creado: {LOCAL_BACKUP_FILE}")

        total_insertados = actualizar_workbook_variables(
            xlsx_path=LOCAL_DOWNLOAD_FILE,
            df_badlar=df_badlar,
            df_tamar=df_tamar,
            df_com=df_com,
        )

        if total_insertados == 0:
            log("Proceso finalizado OK. No había variables nuevas para cargar.")
            return 0

        if DRY_RUN:
            log("DRY_RUN=True. No se subió nada a Drive.")
            return 0

        subir_xlsx_drive(service, DRIVE_FILE_ID, LOCAL_UPDATED_FILE)

        log("Proceso finalizado OK.")
        return 0

    except Exception as e:
        log(f"ERROR: {repr(e)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())