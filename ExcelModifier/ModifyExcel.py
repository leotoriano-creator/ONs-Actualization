# -*- coding: utf-8 -*-
"""
Actualiza Base ONs - Nueva.xlsx directamente como archivo Excel en Google Drive.

Flujo:
    1) Descarga el .xlsx desde Drive usando service account.
    2) Lee DataStorage/a3_on_corporativas.xlsx.
    3) Inserta en hoja BASE solamente filas con fechaLiquidacion estrictamente mayor
       a la última Fecha de Emisión y Liquidación existente en BASE.
    4) Copia formato y fórmulas O:P desde la última fila real.
    5) Para tasa variable:
        - H = Variable
        - L = fórmula de Tasa Base según M: TAMAR/BADLAR
        - M = TAMAR/BADLAR
        - N = Margen
        - J = L + N
    6) Sube el .xlsx actualizado al mismo archivo de Drive.

Requisitos:
    pip install pandas openpyxl google-api-python-client google-auth
"""

from __future__ import annotations

import io
import os
import re
import sys
import json
import base64
import math
import shutil
from copy import copy
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload


# ============================================================
# CONFIG
# ============================================================

DRIVE_FILE_ID = os.getenv("DRIVE_FILE_ID", "19Hr6IBST72-D9j2mNTgcBH_-XmI2F9Py")
WORKSHEET_NAME = "BASE"

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name.lower() == "getdata" else Path.cwd()

SERVICE_ACCOUNT_FILE = PROJECT_ROOT / "ServiceAccount" / "service_account.json"
A3_FILE = PROJECT_ROOT / "DataStorage" / "a3_on_corporativas.xlsx"

LOCAL_DOWNLOAD_FILE = PROJECT_ROOT / "DataStorage" / "Base ONs - Nueva_DESCARGADA.xlsx"
LOCAL_BACKUP_FILE = PROJECT_ROOT / "DataStorage" / f"backup_Base_ONs_Nueva_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
LOCAL_UPDATED_FILE = PROJECT_ROOT / "DataStorage" / "Base ONs - Nueva_ACTUALIZADA.xlsx"
PREVIEW_FILE = PROJECT_ROOT / "DataStorage" / "preview_base_ons_a_insertar.xlsx"

DRY_RUN = os.getenv("DRY_RUN", "false").strip().lower() in {"1", "true", "yes", "y", "si", "sí"}


# ============================================================
# LOG
# ============================================================

def log(msg: str) -> None:
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{ahora} | {msg}")



# ============================================================
# CREDENCIALES GOOGLE
# ============================================================

def cargar_credenciales_google(scopes):
    """
    Railway:
        Lee GOOGLE_SERVICE_ACCOUNT_JSON desde variables de entorno.

    Local:
        Si no existe la variable, usa ServiceAccount/service_account.json.
    """
    raw_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()

    if raw_json:
        try:
            info = json.loads(raw_json)
        except json.JSONDecodeError:
            try:
                decoded = base64.b64decode(raw_json).decode("utf-8")
                info = json.loads(decoded)
            except Exception as e:
                raise ValueError(
                    "GOOGLE_SERVICE_ACCOUNT_JSON existe, pero no pude parsearla "
                    "como JSON directo ni como base64(JSON). Revisá cómo la pegaste en Railway."
                ) from e

        if "private_key" in info and isinstance(info["private_key"], str):
            info["private_key"] = info["private_key"].replace("\\n", "\n")

        return Credentials.from_service_account_info(info, scopes=scopes)

    if SERVICE_ACCOUNT_FILE.exists():
        return Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)

    raise FileNotFoundError(
        "No encontré credenciales de Google. En Railway definí GOOGLE_SERVICE_ACCOUNT_JSON "
        "con el JSON completo de la service account."
    )


# ============================================================
# DRIVE API
# ============================================================

def conectar_drive():
    scopes = ["https://www.googleapis.com/auth/drive"]

    creds = cargar_credenciales_google(scopes)

    try:
        log(f"Service account usada: {creds.service_account_email}")
    except Exception:
        log("Service account usada: desconocida")

    service = build("drive", "v3", credentials=creds)
    return service


def descargar_xlsx_drive(service, file_id: str, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)

    meta = service.files().get(
        fileId=file_id,
        fields="id,name,mimeType",
        supportsAllDrives=True,
    ).execute()

    log(f"Archivo Drive encontrado: {meta.get('name')}")
    log(f"MIME type: {meta.get('mimeType')}")

    mime_type = meta.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.spreadsheet":
        raise RuntimeError(
            "El archivo es Google Sheet nativo. Este script está pensado para editar un .xlsx puro."
        )

    request = service.files().get_media(
        fileId=file_id,
        supportsAllDrives=True,
    )

    fh = io.FileIO(destino, "wb")
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            log(f"Descarga: {int(status.progress() * 100)}%")

    fh.close()

    log(f"Archivo descargado en: {destino}")


def subir_xlsx_drive(service, file_id: str, archivo: Path) -> None:
    media = MediaFileUpload(
        str(archivo),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )

    updated = service.files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()

    log(f"Archivo actualizado en Drive OK. ID: {updated.get('id')}")


# ============================================================
# HELPERS
# ============================================================

def normalizar_texto(x) -> str:
    if x is None:
        return ""

    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass

    s = str(x).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def texto_upper(x) -> str:
    return normalizar_texto(x).upper()


def parse_fecha_a3(x):
    s = normalizar_texto(x)

    if not s or s.startswith("0001-01-01"):
        return pd.NaT

    dt = pd.to_datetime(s, errors="coerce")

    if pd.isna(dt):
        return pd.NaT

    return dt.normalize()


def parse_fecha_excel(x):
    if x is None:
        return pd.NaT

    if isinstance(x, datetime):
        return pd.Timestamp(x).normalize()

    s = str(x).strip()
    if not s:
        return pd.NaT

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)

    if pd.isna(dt):
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)

    if pd.isna(dt):
        return pd.NaT

    return dt.normalize()


def numero_para_excel(x):
    if x is None:
        return None

    try:
        if pd.isna(x):
            return None
    except Exception:
        pass

    try:
        return float(x)
    except Exception:
        return None


def pct_a_decimal(pct):
    """
    Excel guarda porcentajes como decimal:
        100%  -> 1
        5.25% -> 0.0525
    """
    if pct is None or pct == "":
        return None

    try:
        if pd.isna(pct):
            return None
    except Exception:
        pass

    try:
        val = float(pct)
    except Exception:
        return None

    if not math.isfinite(val):
        return None

    return val / 100.0


def parse_numero_es_ar(token: str, contexto: str = ""):
    """
    Parser robusto para valores A3.

    Casos:
        5.25       -> 5.25
        2,99       -> 2.99
        981.50     -> 981.50
        1.016.66   -> 1016.66
        1.071      -> 1071 si es precio por VN 1.000
        0.977275   -> 0.977275
    """
    s = normalizar_texto(token)
    if not s:
        return None

    s = re.sub(r"[^\d,.\-+]", "", s)

    if not s or s in {".", ",", "-", "+"}:
        return None

    contexto = contexto.lower()

    if "." in s and "," in s:
        last_dot = s.rfind(".")
        last_comma = s.rfind(",")

        if last_comma > last_dot:
            s2 = s.replace(".", "").replace(",", ".")
        else:
            s2 = s.replace(",", "")

        try:
            return float(s2)
        except Exception:
            return None

    if "," in s and "." not in s:
        parts = s.split(",")

        if len(parts) == 2 and len(parts[1]) == 3 and contexto == "precio_por_1000":
            s2 = "".join(parts)
        else:
            s2 = s.replace(",", ".")

        try:
            return float(s2)
        except Exception:
            return None

    if "." in s and "," not in s:
        parts = s.split(".")

        if len(parts) > 2:
            s2 = "".join(parts[:-1]) + "." + parts[-1]
            try:
                return float(s2)
            except Exception:
                return None

        entero, dec = parts[0], parts[1]

        if entero in {"0", "-0", "+0"}:
            try:
                return float(s)
            except Exception:
                return None

        if contexto == "precio_por_1000" and len(dec) == 3 and abs(float(entero)) < 10:
            try:
                return float(entero + dec)
            except Exception:
                return None

        try:
            return float(s)
        except Exception:
            return None

    try:
        return float(s)
    except Exception:
        return None


def extraer_numeros(texto: str, contexto: str = "") -> list[float]:
    s = normalizar_texto(texto)

    if not s:
        return []

    tokens = re.findall(r"[-+]?\d[\d.,]*", s)
    nums = []

    for tok in tokens:
        n = parse_numero_es_ar(tok, contexto=contexto)
        if n is not None:
            nums.append(n)

    return nums


# ============================================================
# PARSERS FINANCIEROS
# ============================================================

def valor_corte_contiene_precio(valor_corte: str, variable_licitar: str) -> bool:
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)
    return "PRECIO" in vc or "PRECIO" in vl


def valor_corte_contiene_tasa(valor_corte: str, variable_licitar: str) -> bool:
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)
    return "TASA" in vc or vl in {"TASA", "TEM"}


def valor_corte_contiene_margen(valor_corte: str, variable_licitar: str) -> bool:
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)
    return "MARGEN" in vc or "MARGEN" in vl


def calcular_precio_corte_pct(valor_corte: str, variable_licitar: str):
    """
    Devuelve precio de corte en puntos porcentuales:
        Licita por tasa/margen sin precio explícito -> 100
        USD 981.50 por cada VN 1000 -> 98.15
        USD 1016.66 por cada VN 1000 -> 101.666
        $ 1.071 por cada VN 1000 -> 107.1
        0.977275 -> 97.7275
    """
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)

    if "DESIERTA" in vc:
        return ""

    if not valor_corte_contiene_precio(vc, vl):
        return 100.0

    contexto = "precio_por_1000" if "CADA VNO" in vc or "VN 1.000" in vc or "VNO" in vc else "precio"
    nums = extraer_numeros(valor_corte, contexto=contexto)

    if not nums:
        return ""

    n = nums[-1]

    if contexto == "precio_por_1000":
        return n / 10.0

    if 0 < n <= 2:
        return n * 100.0

    if 2 < n <= 300:
        return n

    return n


def calcular_tna_inicial_pct(valor_corte: str, variable_licitar: str):
    """
    Para tasa fija, devuelve TNA inicial en puntos porcentuales.
    Para margen, devuelve vacío porque se calcula después como Tasa Base + Margen.
    """
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)

    if "DESIERTA" in vc:
        return ""

    if valor_corte_contiene_margen(vc, vl):
        return ""

    if not valor_corte_contiene_tasa(vc, vl):
        return ""

    nums = extraer_numeros(valor_corte, contexto="tasa")

    if not nums:
        return ""

    return nums[-1]


def calcular_margen_pct(valor_corte: str, variable_licitar: str):
    vc = texto_upper(valor_corte)
    vl = texto_upper(variable_licitar)

    if "DESIERTA" in vc:
        return ""

    if not valor_corte_contiene_margen(vc, vl):
        return ""

    nums = extraer_numeros(valor_corte, contexto="tasa")

    if not nums:
        return ""

    return nums[-1]


def moneda_base(moneda: str) -> str:
    m = texto_upper(moneda)

    if "UVA" in m:
        return "UVA"

    if "USD LINK" in m or "DOLAR LINK" in m or "DÓLAR LINK" in m:
        return "Dólar Linked"

    if m == "USD" or "DOLAR" in m or "DÓLAR" in m:
        return "USD"

    if "ARS" in m or "PESO" in m or m == "$":
        return "$"

    return normalizar_texto(moneda)


def regimen_emision_base(tipo: str) -> str:
    t = normalizar_texto(tipo)
    t = re.sub(r"^\s*Privada\s*-\s*", "", t, flags=re.IGNORECASE)
    t = re.sub(r"^\s*Pública\s*-\s*", "", t, flags=re.IGNORECASE)
    return t.strip()


# ============================================================
# FORMATO DE TEXTO
# ============================================================

def formatear_serie_clase(texto: str) -> str:
    """
    Convierte:
        CLASE 32 SERIE I -> Clase 32 Serie I
        SERIE I -> Serie I
        CLASE XXIV -> Clase XXIV
        Clase 5 -> Clase 5
    """
    s = normalizar_texto(texto)

    if not s:
        return ""

    partes = s.split()
    salida = []

    for p in partes:
        up = p.upper()

        if up == "CLASE":
            salida.append("Clase")
        elif up == "SERIE":
            salida.append("Serie")
        else:
            salida.append(up)

    return " ".join(salida)


def extraer_serie_clase(titulo: str) -> str:
    """
    Toma todo lo que viene desde la primera aparición de CLASE o SERIE hasta la derecha.
    """
    t_original = normalizar_texto(titulo)
    t_upper = t_original.upper()

    posiciones = []
    for palabra in ["CLASE", "SERIE"]:
        m = re.search(rf"\b{palabra}\b", t_upper)
        if m:
            posiciones.append(m.start())

    if not posiciones:
        return t_original

    pos = min(posiciones)
    serie = t_original[pos:].strip()

    return formatear_serie_clase(serie)


def formatear_token_sociedad(token: str, idx: int) -> str:
    """
    Formatea una palabra del nombre de la sociedad.
    Conserva siglas y formas societarias.
    """
    if not token:
        return ""

    token = token.strip()
    up = token.upper()

    conectores_minuscula = {
        "DE", "DEL", "LA", "LAS", "LOS", "Y", "E", "EN", "EL", "DA", "DO", "DOS", "DAS"
    }

    siglas_mantener = {
        "S.A.", "S.A", "SA",
        "S.A.U.", "S.A.U", "SAU",
        "S.R.L.", "S.R.L", "SRL",
        "S.G.R.", "S.G.R", "SGR",
        "S.C.A.", "S.C.A", "SCA",
        "S.A.I.C.", "S.A.I.C", "SAIC",
        "S.A.I.C.F.I.A.", "S.A.I.C.F.I.A", "SAICFIA",
        "B.V.", "B.V", "BV",
        "LLC", "INC", "PLC",
        "ON", "VCP", "PYME", "CNV",
        "YPF", "PAMPA", "ICBC", "BBVA", "HSBC", "CMA", "CGC",
        "YPF.", "ICBC.", "BBVA.",
    }

    equivalencias = {
        "SA": "S.A.",
        "S.A": "S.A.",
        "S.A.": "S.A.",
        "SAU": "S.A.U.",
        "S.A.U": "S.A.U.",
        "S.A.U.": "S.A.U.",
        "SRL": "S.R.L.",
        "S.R.L": "S.R.L.",
        "S.R.L.": "S.R.L.",
        "SGR": "S.G.R.",
        "S.G.R": "S.G.R.",
        "S.G.R.": "S.G.R.",
        "SCA": "S.C.A.",
        "S.C.A": "S.C.A.",
        "S.C.A.": "S.C.A.",
        "SAIC": "S.A.I.C.",
        "S.A.I.C": "S.A.I.C.",
        "S.A.I.C.": "S.A.I.C.",
        "SAICFIA": "S.A.I.C.F.I.A.",
        "S.A.I.C.F.I.A": "S.A.I.C.F.I.A.",
        "S.A.I.C.F.I.A.": "S.A.I.C.F.I.A.",
        "BV": "B.V.",
        "B.V": "B.V.",
        "B.V.": "B.V.",
    }

    # Si viene con puntuación final tipo coma, separamos y reponemos.
    sufijo = ""
    while up and up[-1] in [",", ";", ":"]:
        sufijo = up[-1] + sufijo
        token = token[:-1]
        up = token.upper()

    if up in equivalencias:
        return equivalencias[up] + sufijo

    if up in siglas_mantener:
        return up + sufijo

    if idx > 0 and up in conectores_minuscula:
        return up.lower() + sufijo

    # Respeta palabras con guión.
    if "-" in token:
        partes = token.split("-")
        return "-".join(
            p[:1].upper() + p[1:].lower() if p else p
            for p in partes
        ) + sufijo

    return token[:1].upper() + token[1:].lower() + sufijo


def formatear_sociedad(nombre: str) -> str:
    """
    Convierte nombres tipo:
        BANCO DE GALICIA Y BUENOS AIRES S.A.U.
    a:
        Banco de Galicia y Buenos Aires S.A.U.

    Es deliberadamente conservador: mantiene siglas, formas societarias y conectores.
    """
    s = normalizar_texto(nombre)

    if not s:
        return ""

    # Limpieza de espacios raros.
    s = re.sub(r"\s+", " ", s).strip()

    palabras = s.split()
    salida = [formatear_token_sociedad(p, i) for i, p in enumerate(palabras)]

    resultado = " ".join(salida)

    # Correcciones puntuales usuales.
    reemplazos = {
        "Icbc": "ICBC",
        "Bbva": "BBVA",
        "Hsbc": "HSBC",
        "Ypf": "YPF",
        "Pampa": "Pampa",
        "Cgc": "CGC",
        "Sacifia": "S.A.C.I.F.I.A.",
        "Saicfia": "S.A.I.C.F.I.A.",
        "Saic": "S.A.I.C.",
        "S.a.": "S.A.",
        "S.a.u.": "S.A.U.",
        "S.r.l.": "S.R.L.",
        "S.g.r.": "S.G.R.",
        "S.c.a.": "S.C.A.",
    }

    for viejo, nuevo in reemplazos.items():
        resultado = resultado.replace(viejo, nuevo)

    return resultado


# ============================================================
# TASA / PLAZO
# ============================================================

def tasa_base_desde_texto(titulo: str, observaciones: str, variable_licitar: str, valor_corte: str = "") -> str:
    t = (
        f"{normalizar_texto(titulo)} "
        f"{normalizar_texto(observaciones)} "
        f"{normalizar_texto(variable_licitar)} "
        f"{normalizar_texto(valor_corte)}"
    ).upper()

    if "BADLAR" in t:
        return "BADLAR"

    if "TAMAR" in t:
        return "TAMAR"

    if "MARGEN" in t:
        return "TAMAR"

    if "CER" in t:
        return "CER"

    if "UVA" in t:
        return "UVA"

    if "DOLAR LINK" in t or "DÓLAR LINK" in t or "USD LINK" in t:
        return "DLK"

    return ""


def tipo_tasa_fija_variable(variable_licitar: str, valor_corte: str, tasa_base: str) -> str:
    """
    Columna H - Tasa: Fija / Variable.
    """
    vl = texto_upper(variable_licitar)
    vc = texto_upper(valor_corte)

    if "MARGEN" in vl or "MARGEN" in vc:
        return "Variable"

    if tasa_base in {"TAMAR", "BADLAR"}:
        return "Variable"

    if "TASA" in vl or "TASA" in vc or vl == "TEM":
        return "Fija"

    return ""


def tipo_tasa_base_variable(titulo: str, observaciones: str, variable_licitar: str, valor_corte: str) -> str:
    """
    Columna M - Tipo de Tasa.
    Para variables, usamos TAMAR por default salvo que diga BADLAR.
    Para fijas, usamos Fija.
    """
    texto = (
        f"{normalizar_texto(titulo)} "
        f"{normalizar_texto(observaciones)} "
        f"{normalizar_texto(variable_licitar)} "
        f"{normalizar_texto(valor_corte)}"
    ).upper()

    if "MARGEN" in texto:
        if "BADLAR" in texto:
            return "BADLAR"
        return "TAMAR"

    if "BADLAR" in texto:
        return "BADLAR"

    if "TAMAR" in texto:
        return "TAMAR"

    if "TASA" in texto:
        return "Fija"

    return ""


def plazo_meses_desde_fechas(fecha_liq, fecha_vto):
    liq = pd.to_datetime(fecha_liq, errors="coerce")
    vto = pd.to_datetime(fecha_vto, errors="coerce")

    if pd.isna(liq) or pd.isna(vto):
        return None

    if vto <= liq:
        return None

    meses_enteros = (vto.year - liq.year) * 12 + (vto.month - liq.month)
    dias_extra = vto.day - liq.day
    meses = meses_enteros + dias_extra / 30.4375

    if meses <= 0:
        return None

    return int(round(meses, 0))


# ============================================================
# TRANSFORMACIÓN
# ============================================================

def transformar_a_base(df: pd.DataFrame) -> pd.DataFrame:
    needed = [
        "fechaInicio",
        "fechaLiquidacion",
        "fechaVencimiento",
        "titulo",
        "emisor",
        "tipo",
        "moneda",
        "monto_Adjudicado",
        "variableLicitar",
        "valor_Corte",
        "observaciones",
    ]

    for col in needed:
        if col not in df.columns:
            df[col] = pd.NA

    rows = []

    for _, r in df.iterrows():
        fecha_colocacion = parse_fecha_a3(r["fechaInicio"])
        fecha_liq = parse_fecha_a3(r["fechaLiquidacion"])
        fecha_vto = parse_fecha_a3(r["fechaVencimiento"])

        titulo = normalizar_texto(r["titulo"])
        emisor = normalizar_texto(r["emisor"])
        variable = normalizar_texto(r["variableLicitar"])
        valor_corte = normalizar_texto(r["valor_Corte"])
        observaciones = normalizar_texto(r["observaciones"])

        tasa_base_detectada = tasa_base_desde_texto(
            titulo=titulo,
            observaciones=observaciones,
            variable_licitar=variable,
            valor_corte=valor_corte,
        )

        tasa_fija_variable = tipo_tasa_fija_variable(
            variable_licitar=variable,
            valor_corte=valor_corte,
            tasa_base=tasa_base_detectada,
        )

        tipo_tasa_base = tipo_tasa_base_variable(
            titulo=titulo,
            observaciones=observaciones,
            variable_licitar=variable,
            valor_corte=valor_corte,
        )

        precio_corte_pct = calcular_precio_corte_pct(valor_corte, variable)
        tna_inicial_pct = calcular_tna_inicial_pct(valor_corte, variable)
        margen_pct = calcular_margen_pct(valor_corte, variable)
        plazo_meses = plazo_meses_desde_fechas(fecha_liq, fecha_vto)

        if tasa_fija_variable == "Variable":
            tna_inicial_valor = None
            tasa_base_valor = None
            if not tipo_tasa_base:
                tipo_tasa_base = "TAMAR"
        else:
            tna_inicial_valor = pct_a_decimal(tna_inicial_pct)
            tasa_base_valor = None

        rows.append({
            "Fecha Colocación": fecha_colocacion,
            "Fecha de Emisión y Liquidación": fecha_liq,
            "Sociedad": formatear_sociedad(emisor),
            "Serie/Clase": extraer_serie_clase(titulo),
            "Moneda": moneda_base(r["moneda"]),
            "Monto nominal (moneda emisión)": numero_para_excel(r["monto_Adjudicado"]),
            "Régimen de emisión": regimen_emision_base(r["tipo"]),
            "Tasa": tasa_fija_variable,
            "Precio de corte": pct_a_decimal(precio_corte_pct),
            "TNA inicial": tna_inicial_valor,
            "Plazo (meses)": plazo_meses,
            "Tasa Base": tasa_base_valor,
            "Tipo de Tasa": tipo_tasa_base,
            "Margen": pct_a_decimal(margen_pct),
        })

    columnas = [
        "Fecha Colocación",
        "Fecha de Emisión y Liquidación",
        "Sociedad",
        "Serie/Clase",
        "Moneda",
        "Monto nominal (moneda emisión)",
        "Régimen de emisión",
        "Tasa",
        "Precio de corte",
        "TNA inicial",
        "Plazo (meses)",
        "Tasa Base",
        "Tipo de Tasa",
        "Margen",
    ]

    return pd.DataFrame(rows)[columnas]


# ============================================================
# EXCEL WRITE
# ============================================================

def ultima_fila_real_ws(ws) -> int:
    ultima = 1

    for row in range(2, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        c = ws.cell(row=row, column=3).value

        if a not in [None, ""] or b not in [None, ""] or c not in [None, ""]:
            ultima = row

    return ultima


def ultima_fecha_liquidacion_ws(ws, ultima_fila: int):
    fechas = []

    for row in range(2, ultima_fila + 1):
        val = ws.cell(row=row, column=2).value
        dt = parse_fecha_excel(val)

        if not pd.isna(dt):
            fechas.append(dt)

    if not fechas:
        return pd.NaT

    return max(fechas)


def copiar_estilo_celda(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def traducir_formula(formula: str, origen: str, destino: str):
    try:
        return Translator(formula, origin=origen).translate_formula(destino)
    except Exception:
        return formula


def aplicar_alineacion_nuevas_filas(ws, fila_inicio: int, fila_fin: int) -> None:
    """
    Ajusta estética de filas nuevas.
    """
    for row in range(fila_inicio, fila_fin + 1):
        # Fechas centradas.
        for col in [1, 2]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Sociedad alineada a izquierda, como la base vieja.
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")

        # Serie/Clase centrada.
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")

        # Moneda / Monto / Régimen.
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Tasa, precio, TNA, plazo, tasa base, tipo tasa, margen.
        for col in [8, 9, 10, 11, 12, 13, 14]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # COM3500 y MM USD.
        for col in [15, 16]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")


def insertar_en_workbook(xlsx_path: Path, df_base: pd.DataFrame) -> None:
    wb = load_workbook(xlsx_path)

    if WORKSHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja {WORKSHEET_NAME} en el archivo.")

    ws = wb[WORKSHEET_NAME]

    ultima_fila = ultima_fila_real_ws(ws)
    log(f"Última fila real en BASE: {ultima_fila}")

    fila_inicio = ultima_fila + 1
    fila_fin = fila_inicio + len(df_base) - 1

    log(f"Insertando filas: {len(df_base)}")
    log(f"Rango destino: A{fila_inicio}:N{fila_fin}")

    for i, row_data in enumerate(df_base.itertuples(index=False), start=fila_inicio):
        # Copia formato A:P desde última fila real.
        for col in range(1, 17):
            src = ws.cell(row=ultima_fila, column=col)
            dst = ws.cell(row=i, column=col)
            copiar_estilo_celda(src, dst)

        # Valores A:N.
        values = list(row_data)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.value = value

        # Si la tasa es variable:
        # L = fórmula de tasa base.
        # M = TAMAR/BADLAR.
        # J = L + N.
        tasa_col_h = ws.cell(row=i, column=8).value
        tipo_tasa_col_m = ws.cell(row=i, column=13).value

        if str(tasa_col_h).strip().upper() == "VARIABLE":
            if not tipo_tasa_col_m:
                ws.cell(row=i, column=13).value = "TAMAR"

            ws.cell(row=i, column=12).value = (
                f'=+SI(M{i}="BADLAR",'
                f'BUSCARV(BASE!B{i},VARIABLES!$A:$B,2,FALSO)/100,'
                f'BUSCARV(BASE!B{i},VARIABLES!$D:$E,2,FALSO)/100)'
            )

            ws.cell(row=i, column=10).value = f"=L{i}+N{i}"

        else:
            ws.cell(row=i, column=12).value = ""

        # Fórmulas O:P desde última fila real, traducidas a la nueva fila.
        for col in range(15, 17):
            src = ws.cell(row=ultima_fila, column=col)
            dst = ws.cell(row=i, column=col)

            if isinstance(src.value, str) and src.value.startswith("="):
                dst.value = traducir_formula(src.value, src.coordinate, dst.coordinate)
            else:
                dst.value = src.value

    # Formatos explícitos.
    for row in range(fila_inicio, fila_fin + 1):
        ws.cell(row=row, column=1).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=2).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=9).number_format = "0.00%"
        ws.cell(row=row, column=10).number_format = "0.00%"
        ws.cell(row=row, column=12).number_format = "0.00%"
        ws.cell(row=row, column=14).number_format = "0.00%"

    aplicar_alineacion_nuevas_filas(ws, fila_inicio, fila_fin)

    wb.save(LOCAL_UPDATED_FILE)
    log(f"Workbook actualizado guardado en: {LOCAL_UPDATED_FILE}")


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    try:
        if not A3_FILE.exists():
            raise FileNotFoundError(f"No encontré el archivo scrapeado: {A3_FILE}")

        service = conectar_drive()

        descargar_xlsx_drive(service, DRIVE_FILE_ID, LOCAL_DOWNLOAD_FILE)
        shutil.copy2(LOCAL_DOWNLOAD_FILE, LOCAL_BACKUP_FILE)
        log(f"Backup local creado: {LOCAL_BACKUP_FILE}")

        log(f"Leyendo archivo A3: {A3_FILE}")
        df_a3 = pd.read_excel(A3_FILE)

        if df_a3.empty:
            log("El archivo A3 está vacío. No hay nada para insertar.")
            return 0

        if "fechaLiquidacion" not in df_a3.columns:
            raise ValueError("El archivo A3 no tiene columna fechaLiquidacion.")

        if "valor_Corte" in df_a3.columns:
            antes = len(df_a3)
            df_a3 = df_a3[
                ~df_a3["valor_Corte"].astype(str).str.upper().str.contains("DESIERTA", na=False)
            ].copy()
            log(f"Filas descartadas por DESIERTA: {antes - len(df_a3)}")

        if "monto_Adjudicado" in df_a3.columns:
            antes = len(df_a3)
            df_a3["monto_Adjudicado_num"] = pd.to_numeric(df_a3["monto_Adjudicado"], errors="coerce")
            df_a3 = df_a3[df_a3["monto_Adjudicado_num"].fillna(0) > 0].copy()
            log(f"Filas descartadas por monto_Adjudicado <= 0: {antes - len(df_a3)}")

        df_a3["fechaLiquidacion_dt"] = df_a3["fechaLiquidacion"].apply(parse_fecha_a3)

        if "fechaInicio" in df_a3.columns:
            df_a3["fechaInicio_dt"] = df_a3["fechaInicio"].apply(parse_fecha_a3)
        else:
            df_a3["fechaInicio_dt"] = pd.NaT

        wb_check = load_workbook(LOCAL_DOWNLOAD_FILE, data_only=False)
        ws_check = wb_check[WORKSHEET_NAME]

        ultima_fila = ultima_fila_real_ws(ws_check)
        ultima_fecha = ultima_fecha_liquidacion_ws(ws_check, ultima_fila)

        if pd.isna(ultima_fecha):
            log("No encontré última fecha válida en BASE. Se insertará todo.")
            df_nuevo = df_a3.copy()
        else:
            log(f"Última Fecha de Emisión y Liquidación en BASE: {ultima_fecha.strftime('%Y-%m-%d')}")
            df_nuevo = df_a3[df_a3["fechaLiquidacion_dt"] > ultima_fecha].copy()

        df_nuevo = df_nuevo.sort_values(
            by=["fechaLiquidacion_dt", "fechaInicio_dt", "emisor", "titulo"],
            ascending=[True, True, True, True],
            na_position="last",
        ).reset_index(drop=True)

        if df_nuevo.empty:
            log("No hay filas nuevas para insertar. Proceso finalizado OK.")
            return 0

        df_base = transformar_a_base(df_nuevo)

        PREVIEW_FILE.parent.mkdir(parents=True, exist_ok=True)
        df_base.to_excel(PREVIEW_FILE, index=False)
        log(f"Preview generado: {PREVIEW_FILE}")

        if DRY_RUN:
            log("DRY_RUN=True. No se escribió nada en Drive.")
            return 0

        insertar_en_workbook(LOCAL_DOWNLOAD_FILE, df_base)
        subir_xlsx_drive(service, DRIVE_FILE_ID, LOCAL_UPDATED_FILE)

        log("Proceso finalizado OK.")
        return 0

    except Exception as e:
        log(f"ERROR: {repr(e)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())